"""
XLSX Mapper - Maps XLSX templates to UnderwritingRequest domain model.
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Optional

from ..types import UnderwritingRequest
from .json_mapper import JSONMapper


@dataclass
class XLSXMappingResult:
    """Result of XLSX mapping."""

    request: Optional[UnderwritingRequest] = None
    warnings: list = field(default_factory=list)
    errors: list = field(default_factory=list)


class XLSXMapper:
    """Maps XLSX underwriting templates to canonical UnderwritingRequest."""

    def __init__(self):
        self.json_mapper = JSONMapper()

    def map_to_request(
        self,
        file_path: Path,
        template_type: str = "standard",
        request_id: Optional[str] = None,
    ) -> XLSXMappingResult:
        """
        Map XLSX template to UnderwritingRequest.

        Args:
            file_path: Path to XLSX file
            template_type: Template format identifier
            request_id: Optional request ID

        Returns:
            XLSXMappingResult
        """
        result = XLSXMappingResult()

        try:
            # Try to import openpyxl
            try:
                from openpyxl import load_workbook
            except ImportError:
                result.errors.append("openpyxl required for XLSX ingestion: pip install openpyxl")
                return result

            # Load workbook
            wb = load_workbook(file_path, data_only=True)

            # Parse based on template type
            if template_type == "standard":
                data = self._parse_standard_template(wb)
            else:
                result.errors.append(f"Unknown template type: {template_type}")
                return result

            # Map to request using JSON mapper
            json_result = self.json_mapper.map_to_request(
                data,
                request_id=request_id,
            )

            result.request = json_result.request
            result.warnings = json_result.warnings
            result.errors = json_result.errors

        except (ValueError, TypeError, AttributeError, KeyError, OSError) as e:
            result.errors.append(f"XLSX mapping error: {str(e)}")

        return result

    def _parse_standard_template(self, wb) -> Dict[str, Any]:
        """Parse standard underwriting template format."""
        data = {}

        # Sheet mapping
        sheet_mapping = {
            "Borrower": "borrower",
            "Financials": "financials",
            "Collateral": "collateral",
            "Credit": "credit",
            "Banking": "banking",
            "Request": "request_info",
        }

        for sheet_name, data_key in sheet_mapping.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = self._parse_sheet(ws)
                if data_key == "request_info":
                    # Merge request info into root
                    data.update(sheet_data)
                else:
                    data[data_key] = sheet_data

        return data

    def _parse_sheet(self, ws) -> Dict[str, Any]:
        """Parse a worksheet into dict."""
        data = {}

        # Assume two-column format: Field | Value
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
            if len(row) >= 2 and row[0].value:
                field_name = str(row[0].value).strip().lower().replace(" ", "_")
                value = row[1].value
                data[field_name] = value

        return data
