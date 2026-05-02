"""Fort Knox v2 — XLSX READ-ONLY exporter.

DOES NOT compute status. DOES NOT override status. DOES NOT read any
authoritative source except the compiler's JSON report.

If the JSON report is missing, this script fails. It will not fabricate.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = REPO_ROOT / "artifacts" / "certification"
REPORT_PATH = OUTPUT_DIR / "final_requirement_signoff_report.json"
MERKLE_PATH = OUTPUT_DIR / "final_requirement_signoff_report.merkle.json"
SIG_PATH = OUTPUT_DIR / "final_requirement_signoff_report.signature.json"
OUT_XLSX = OUTPUT_DIR / "final_requirement_signoff_report.xlsx"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_FILL = {
    "SIGNED_OFF":   PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    "BLOCKED":      PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
    "NOT_VERIFIED": PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
}


def main() -> int:
    if not REPORT_PATH.exists():
        print(f"FATAL: compiler report missing at {REPORT_PATH.relative_to(REPO_ROOT)}. "
              f"Run scripts/compile_requirement_signoff.py first.", file=sys.stderr)
        return 2
    report = json.loads(REPORT_PATH.read_text(encoding="utf-8"))
    merkle = json.loads(MERKLE_PATH.read_text(encoding="utf-8")) if MERKLE_PATH.exists() else None
    sig = json.loads(SIG_PATH.read_text(encoding="utf-8")) if SIG_PATH.exists() else None
    s = report["summary"]

    wb = openpyxl.Workbook()

    # --- READ-ONLY banner + provenance -------------------------------------
    ws = wb.active
    ws.title = "Read-Only View"
    ws.append(["⚠️ READ-ONLY VIEW. JSON compiler report is the authority."])
    ws.append(["Manual edits here do NOT affect certification status."])
    ws.append([])
    ws["A1"].font = Font(bold=True, color="B71C1C", size=13)
    rows = [
        ("Trust level", report.get("trust_level", "?")),
        ("Run timestamp (UTC)", report.get("run_timestamp_utc", "?")),
        ("Compiler version", report.get("compiler_version", "?")),
        ("Compiler sha256", report.get("compiler_sha256", "")),
        ("Git commit", report.get("git_commit", "?")),
        ("Git dirty", str(report.get("git_dirty", False))),
        ("Requirements source SHA256", report.get("requirements_source_sha256", "")),
        ("Evidence assertions SHA256", report.get("evidence_assertions_sha256", "")),
        ("Row digest", report.get("row_digest", "")),
        ("Evidence digest", report.get("evidence_digest", "")),
        ("Merkle root", merkle["root"] if merkle else "(missing)"),
        ("Merkle leaf count", merkle["leaf_count"] if merkle else "(missing)"),
        ("Signature status", sig["signature_verification_status"] if sig else "(missing)"),
        ("Signer identity", sig.get("signer_identity") if sig else "(missing)"),
        ("", ""),
        ("Total", s["total"]),
        ("SIGNED_OFF", f"{s['signed_off']} ({s['percent_signed_off']}%)"),
        ("BLOCKED", s["blocked"]),
        ("NOT_VERIFIED", s["not_verified"]),
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80

    # --- By Claim Type -----------------------------------------------------
    ws2 = wb.create_sheet("By Claim Type")
    headers = ["Claim Type", "Total", "SIGNED_OFF", "BLOCKED", "NOT_VERIFIED"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for ct in sorted(s.get("by_claim_type", {})):
        b = s["by_claim_type"][ct]
        ws2.append([ct, b["total"], b["signed_off"], b["blocked"], b["not_verified"]])
    for i in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 28

    # --- Rows --------------------------------------------------------------
    ws3 = wb.create_sheet("Rows")
    headers = ["req_id", "computed_status", "claim_type", "priority", "requirement_group",
               "is_final_100%", "title", "blocking_gap", "row_digest", "assertion_count"]
    ws3.append(headers)
    for cell in ws3[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in sorted(report["rows"], key=lambda r: r["req_id"]):
        ws3.append([
            row["req_id"], row["computed_status"], row["claim_type"],
            row.get("priority", ""), row.get("requirement_group", ""),
            "YES" if row.get("is_final_hundred_percent_row") else "",
            row.get("title", ""),
            row.get("blocking_gap") or "",
            row.get("row_digest", ""),
            len(row.get("assertion_ids", [])),
        ])
        last = ws3.max_row
        ws3.cell(row=last, column=2).fill = STATUS_FILL.get(
            row["computed_status"], PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))
    widths = [16, 16, 26, 10, 26, 14, 60, 80, 70, 14]
    for i, w in enumerate(widths, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # --- Controls ----------------------------------------------------------
    ws4 = wb.create_sheet("Controls")
    headers = ["req_id", "control", "passed", "reason", "assertion_id", "artifact_path"]
    ws4.append(headers)
    for cell in ws4[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in sorted(report["rows"], key=lambda r: r["req_id"]):
        for c in row["controls"]:
            ws4.append([row["req_id"], c["name"],
                        "PASS" if c["passed"] else "FAIL",
                        c.get("reason") or "",
                        c.get("assertion_id") or "",
                        c.get("artifact_path") or ""])
            last = ws4.max_row
            ws4.cell(row=last, column=3).fill = (
                STATUS_FILL["SIGNED_OFF"] if c["passed"] else STATUS_FILL["BLOCKED"])
    widths = [16, 32, 8, 80, 50, 60]
    for i, w in enumerate(widths, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"

    # --- Provenance tail ---------------------------------------------------
    ws5 = wb.create_sheet("Provenance")
    ws5.append(["READ-ONLY VIEW. JSON compiler report is authority."])
    ws5.append(["Exporter: scripts/export_signoff_to_xlsx.py"])
    ws5.append(["Source: artifacts/certification/final_requirement_signoff_report.json"])
    ws5.append([])
    ws5.append(["This file never computes certification status."])
    ws5.append(["Manual edits are invisible to the compiler and bundle verifier."])
    ws5["A1"].font = Font(bold=True, color="B71C1C")
    ws5.column_dimensions["A"].width = 80

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"[export_signoff_to_xlsx] wrote {OUT_XLSX.relative_to(REPO_ROOT)}")
    print(f"  rollup (read-only, from JSON): "
          f"signed_off={s['signed_off']} blocked={s['blocked']} not_verified={s['not_verified']}")
    print(f"  trust_level: {report.get('trust_level')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
