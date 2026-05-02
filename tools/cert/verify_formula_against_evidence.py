"""F8.1 — Verify the formula's expected output matches CSV signoff_status.

Mirrors the formula logic (in `install_xlsx_formulas.py`) in pure Python
and re-derives `computed_signoff_status` for every row from the
evidence-input cells. Compares against the legacy `signoff_status` column
(col 33). Drift = manual_override_detected.

Output: artifacts/certification/formula_verification_report.json

Exit:
  0 — formula output matches CSV signoff_status for ALL 87 rows
  1 — drift detected (manual_override would fire on N rows)
  2 — harness error
"""

from __future__ import annotations

import csv
import datetime
import json
import sys
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened_FULL_OVERWRITE.xlsx"
)
CSV_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened.csv"
)
MATRIX_PATH = REPO_ROOT / "tools" / "cert" / "required_evidence_matrix.yaml"
REPORT_PATH = REPO_ROOT / "artifacts" / "certification" / "formula_verification_report.json"


def _truthy(v) -> bool:
    if v is True:
        return True
    if isinstance(v, str) and v.strip().lower() == "true":
        return True
    if v == 1:
        return True
    return False


def _evaluate(claim_type: str, evidence: dict, matrix: dict) -> str:
    """Replicate the XLSX formula in pure Python."""
    vs = evidence.get("verifier_status") or ""
    ve = evidence.get("verifier_exit_code")
    lv = evidence.get("last_verified_at_utc") or ""

    # universal_blank: NOT_VERIFIED
    if not vs or not lv:
        return "NOT_VERIFIED"
    if vs == "NOT_VERIFIED":
        return "NOT_VERIFIED"
    if vs == "BLOCKED":
        return "BLOCKED"
    # universal_pass
    if not (vs == "PASS" and ve == 0 and lv):
        return "BLOCKED"
    # claim-type required-evidence
    required = matrix.get("per_claim_type_required", {}).get(claim_type)
    if required is None:
        # unknown claim_type — universal pass alone is enough
        return "SIGNED_OFF"
    for field in required:
        if not _truthy(evidence.get(field)):
            return "BLOCKED"
    return "SIGNED_OFF"


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"FATAL: XLSX not found", file=sys.stderr); return 2
    if not MATRIX_PATH.exists():
        print(f"FATAL: matrix YAML not found", file=sys.stderr); return 2
    matrix = yaml.safe_load(MATRIX_PATH.read_text(encoding="utf-8"))
    evidence_cols = matrix["evidence_input_columns"]

    # Read CSV signoff_status for cross-check
    csv.field_size_limit(min(sys.maxsize, 2**31 - 1))
    with open(CSV_PATH, encoding="utf-8") as f:
        csv_signoff = {r["req_id"]: r["signoff_status"] for r in csv.DictReader(f)}

    # Read XLSX evidence rows
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb["Requirements_Full_Overwrite"]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}

    per_req: list[dict] = []
    drift_rows: list[dict] = []
    rollup_formula = {"SIGNED_OFF": 0, "BLOCKED": 0, "NOT_VERIFIED": 0}
    rollup_csv = {"SIGNED_OFF": 0, "BLOCKED": 0, "NOT_VERIFIED": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = row[col_idx["req_id"]]
        if not rid:
            continue
        ct = row[col_idx["claim_type"]] or ""
        evidence = {c: row[col_idx[c]] for c in evidence_cols if c in col_idx}
        formula_out = _evaluate(ct, evidence, matrix)
        csv_status = csv_signoff.get(rid, "(missing)")
        rollup_formula[formula_out] = rollup_formula.get(formula_out, 0) + 1
        rollup_csv[csv_status] = rollup_csv.get(csv_status, 0) + 1
        entry = {
            "req_id": rid,
            "claim_type": ct,
            "csv_signoff_status": csv_status,
            "formula_computed_signoff_status": formula_out,
            "matches": csv_status == formula_out,
        }
        per_req.append(entry)
        if not entry["matches"]:
            drift_rows.append({**entry, "evidence_snapshot": {
                k: v for k, v in evidence.items() if v not in (None, "", False)
            }})

    overall = "PASS" if not drift_rows else "DRIFT_DETECTED"
    report = {
        "tool": "tools/cert/verify_formula_against_evidence.py",
        "evaluated_at_utc": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        "overall_result": overall,
        "rollup_csv": rollup_csv,
        "rollup_formula": rollup_formula,
        "match_count": len(per_req) - len(drift_rows),
        "drift_count": len(drift_rows),
        "drift_rows": drift_rows[:30],   # first 30 for log size
        "drift_total": len(drift_rows),
        "per_req_summary_count": len(per_req),
    }
    REPORT_PATH.write_text(
        json.dumps(report, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8"
    )

    print(f"[verify_formula] overall={overall}")
    print(f"  rollup_csv     : {rollup_csv}")
    print(f"  rollup_formula : {rollup_formula}")
    print(f"  matches: {len(per_req) - len(drift_rows)}/{len(per_req)}")
    if drift_rows:
        print(f"  DRIFT — {len(drift_rows)} rows where formula and CSV disagree:")
        for d in drift_rows[:10]:
            print(f"    {d['req_id']} [{d['claim_type']}]: csv={d['csv_signoff_status']} formula={d['formula_computed_signoff_status']}")
        if len(drift_rows) > 10:
            print(f"    ... and {len(drift_rows) - 10} more in {REPORT_PATH.relative_to(REPO_ROOT)}")
    print(f"  wrote: {REPORT_PATH.relative_to(REPO_ROOT)}")
    return 0 if not drift_rows else 1


if __name__ == "__main__":
    sys.exit(main())
