"""Sync CSV signoff_status FROM formula output (not the other way around).

Per operator directive 2026-05-01 15:50: a row's status is determined
by formula, not by manual edit. The CSV signoff_status mirrors that
verdict. This helper re-derives the formula in pure Python (mirror of
the XLSX formula in `install_xlsx_formulas.py` / `verify_formula_against_evidence.py`)
and updates the CSV's signoff_status, signoff_evidence_artifact,
signoff_evidence_summary, and signoff_checked_at_utc accordingly.

Atomic CSV write + receipt JSON.
"""
from __future__ import annotations

import csv
import datetime
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened_FULL_OVERWRITE.xlsx"
)
CSV_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened.csv"
)
MATRIX_PATH = REPO_ROOT / "tools" / "cert" / "required_evidence_matrix.yaml"
RECEIPT_DIR = REPO_ROOT / "artifacts" / "certification" / "csv_signoff_updates"


def _truthy(v) -> bool:
    if v is True: return True
    if isinstance(v, str) and v.strip().lower() == "true": return True
    if v == 1: return True
    return False


def _evaluate(claim_type, evidence, matrix):
    vs = evidence.get("verifier_status") or ""
    ve = evidence.get("verifier_exit_code")
    lv = evidence.get("last_verified_at_utc") or ""
    if not vs or not lv: return "NOT_VERIFIED"
    if vs == "NOT_VERIFIED": return "NOT_VERIFIED"
    if vs == "BLOCKED": return "BLOCKED"
    if not (vs == "PASS" and ve == 0 and lv): return "BLOCKED"
    required = matrix.get("per_claim_type_required", {}).get(claim_type)
    if required is None: return "SIGNED_OFF"
    for f in required:
        if not _truthy(evidence.get(f)): return "BLOCKED"
    return "SIGNED_OFF"


def main() -> int:
    started_at = datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")
    matrix = yaml.safe_load(MATRIX_PATH.read_text(encoding="utf-8"))
    evidence_cols = matrix["evidence_input_columns"]

    # Read XLSX
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb["Requirements_Full_Overwrite"]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}

    formula_status_by_req: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = row[col_idx["req_id"]]
        if not rid:
            continue
        ct = row[col_idx["claim_type"]] or ""
        evidence = {c: row[col_idx[c]] for c in evidence_cols if c in col_idx}
        status = _evaluate(ct, evidence, matrix)
        formula_status_by_req[str(rid).strip()] = {
            "formula_status": status,
            "verifier_report_artifact": evidence.get("verifier_report_artifact") or "",
            "verifier_status": evidence.get("verifier_status") or "",
            "last_verified_at_utc": evidence.get("last_verified_at_utc") or "",
            "claim_type": ct,
        }

    # Read CSV
    csv.field_size_limit(min(sys.maxsize, 2**31 - 1))
    with open(CSV_PATH, encoding="utf-8", newline="") as f:
        rdr = csv.DictReader(f)
        fieldnames = rdr.fieldnames
        rows = list(rdr)

    # Apply
    n_changed = 0
    changes_log = []
    for r in rows:
        rid = r["req_id"]
        f = formula_status_by_req.get(rid)
        if not f:
            continue
        new_status = f["formula_status"]
        old_status = r.get("signoff_status", "") or ""
        if old_status != new_status:
            r["signoff_status"] = new_status
            r["signoff_checked_at_utc"] = f["last_verified_at_utc"] or started_at
            # Update artifact / summary based on status transition
            if new_status == "SIGNED_OFF":
                r["signoff_evidence_artifact"] = f["verifier_report_artifact"]
                r["signoff_evidence_summary"] = (
                    f"Formula-derived SIGNED_OFF on {started_at}. "
                    f"claim_type={f['claim_type']}; verifier_status=PASS; backing artifact="
                    f"{f['verifier_report_artifact']}."
                )
            elif new_status == "BLOCKED":
                # Preserve existing summary if present (from F6 emission), else generic
                if not (r.get("signoff_evidence_summary") or "").startswith("Wave"):
                    r["signoff_evidence_summary"] = (
                        f"Formula-derived BLOCKED on {started_at}. "
                        f"verifier_status={f['verifier_status'] or '(empty)'}; required gate(s) not TRUE for claim_type={f['claim_type']}."
                    )
            else:  # NOT_VERIFIED
                if not r.get("signoff_evidence_summary"):
                    r["signoff_evidence_summary"] = f"Formula-derived NOT_VERIFIED on {started_at}."
            n_changed += 1
            changes_log.append({"req_id": rid, "from": old_status, "to": new_status})

    # Atomic CSV write
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="csv_sync_", suffix=".csv", dir=str(CSV_PATH.parent))
    os.close(tmp_fd)
    try:
        with open(tmp_path, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        shutil.move(tmp_path, CSV_PATH)
    except Exception:
        try: os.unlink(tmp_path)
        except OSError: pass
        raise

    # Receipt
    RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
    receipt = {
        "tool": "tools/cert/sync_csv_from_formula.py",
        "executed_at_utc": started_at,
        "csv_path": str(CSV_PATH),
        "n_rows_total": len(rows),
        "n_rows_changed": n_changed,
        "changes": changes_log,
    }
    receipt_path = RECEIPT_DIR / f"{started_at.replace(':', '-')}_csv_from_formula.json"
    receipt_path.write_text(json.dumps(receipt, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    # Rollup
    rollup = {"SIGNED_OFF": 0, "BLOCKED": 0, "NOT_VERIFIED": 0}
    for r in rows:
        rollup[r["signoff_status"]] = rollup.get(r["signoff_status"], 0) + 1

    print(f"[sync_csv_from_formula] changed {n_changed}/{len(rows)} rows")
    print(f"  rollup: {rollup}")
    if changes_log:
        for c in changes_log[:20]:
            print(f"    {c['req_id']}: {c['from']} -> {c['to']}")
        if len(changes_log) > 20:
            print(f"    ... and {len(changes_log) - 20} more")
    print(f"  receipt: {receipt_path.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
