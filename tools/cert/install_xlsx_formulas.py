"""F0.2 — Install formulas in cols 65-68 of the XLSX.

Writes Excel formulas (NOT static values) to:
  - col 65: computed_acceptance_status
  - col 66: computed_signoff_status
  - col 67: computed_blocking_gap
  - col 68: manual_override_detected

The formulas reference:
  - col 7  (G):  claim_type — drives the per-type required-evidence rule
  - cols 37-62 (AK..BJ): evidence-input fields
  - col 33 (AG): legacy signoff_status — used only by manual_override_detected

Per plan `runtime-cert-formula-driven-signoff-a8f5c2.md` §2.3 / F0.2.
SSOT for the per-claim_type rule: tools/cert/required_evidence_matrix.yaml.
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened_FULL_OVERWRITE.xlsx"
)


def _col(headers: list[str], name: str) -> str:
    """Return the Excel column letter for header ``name``."""
    return get_column_letter(headers.index(name) + 1)


def _build_formulas(headers: list[str], row: int) -> dict[str, str]:
    """Build the 4 formulas for a given data row (Excel-1-based row index).

    Per-claim_type required-evidence map mirrors the YAML SSOT exactly.
    Formula structure:

      computed_signoff_status =
        IF(verifier_status<>"PASS" OR verifier_exit_code<>0 OR last_verified_at_utc="",
           IF(verifier_status="BLOCKED", "BLOCKED", "NOT_VERIFIED"),
           IF(<claim_type-specific gate AND chain>, "SIGNED_OFF", "BLOCKED"))

    The claim_type branch is a nested IF chain across the 10 enumerated
    claim_types. An unknown claim_type falls through to the universal
    rule (verifier_status=PASS suffices).
    """

    G = _col(headers, "claim_type")            # claim_type
    AG = _col(headers, "signoff_status")       # legacy mirror
    # Evidence-input columns
    vs = _col(headers, "verifier_status")
    ve = _col(headers, "verifier_exit_code")
    lv = _col(headers, "last_verified_at_utc")
    cig = _col(headers, "ci_gate_verified")
    lay = _col(headers, "layer_boundary_verified")
    rav = _col(headers, "required_artifacts_verified")
    aph = _col(headers, "artifact_payload_hash_verified")
    rev = _col(headers, "runtime_evidence_verified")
    emhv = _col(headers, "evidence_manifest_hash_verified")
    ot = _col(headers, "otel_trace_verified")
    srb = _col(headers, "source_root_binding_verified")
    nob = _col(headers, "no_bypass_verified")
    pev = _col(headers, "positive_evidence_verified")
    rrv = _col(headers, "replay_receipt_verified")
    csv_ = _col(headers, "certifier_signature_verified")
    efv = _col(headers, "expected_fail_reason_verified")
    vra = _col(headers, "verifier_report_artifact")

    R = row  # row index (Excel 1-based)

    # Universal "row even has any input" gate. If verifier_status is
    # blank, the row is genuinely NOT_VERIFIED.
    universal_blank = f'OR({vs}{R}="", {lv}{R}="")'

    # Universal PASS-input gate. Must be PASS+exit-0+timestamped.
    universal_pass = f'AND({vs}{R}="PASS", {ve}{R}=0, {lv}{R}<>"")'

    # Per-claim_type required-evidence boolean (must be TRUE-truthy):
    # Excel truthy: TRUE, "TRUE", "true", 1. We treat any cell whose
    # text-equivalent is TRUE/"TRUE"/1 as truthy via a helper:
    def truthy(cell: str) -> str:
        return f'OR({cell}{R}=TRUE, {cell}{R}="TRUE", {cell}{R}="true", {cell}{R}=1)'

    branches = {
        "MATRIX_GOVERNANCE":            f'AND({truthy(cig)})',
        "STATIC_ENFORCEMENT":           f'AND({truthy(cig)}, {truthy(lay)})',
        "STATIC_CONTRACT":              f'AND({truthy(rav)}, {truthy(aph)})',
        "COMPONENT_RUNTIME":            f'AND({truthy(rev)}, {truthy(emhv)})',
        "INTEGRATED_RUNTIME":           f'AND({truthy(rev)}, {truthy(ot)}, {truthy(srb)}, {truthy(aph)})',
        "NO_BYPASS_RUNTIME":            f'AND({truthy(nob)}, {truthy(rev)})',
        "COMPOSITION_RUNTIME":          f'AND({truthy(rev)}, {truthy(pev)})',
        "OBSERVABILITY_RUNTIME":        f'AND({truthy(ot)})',
        "REPLAY_RUNTIME":               f'AND({truthy(rrv)})',
        "PRODUCTION_DEPENDENCY_RUNTIME": f'AND({truthy(rev)}, {truthy(csv_)})',
    }

    # Build a nested IF chain over claim_type cell value.
    chain_parts = []
    for ct, expr in branches.items():
        chain_parts.append(f'IF({G}{R}="{ct}", {expr}, ')
    # Default branch — unknown claim_type, universal pass alone is enough.
    chain_parts.append("TRUE")
    chain_parts.append(")" * len(branches))
    claim_type_chain = "".join(chain_parts)

    # computed_signoff_status (col 66)
    # Order of evaluation:
    #   1. blank evidence (no vs OR no lv) -> NOT_VERIFIED
    #   2. explicit "NOT_VERIFIED" verifier_status -> NOT_VERIFIED
    #   3. explicit "BLOCKED" verifier_status -> BLOCKED
    #   4. universal_pass (PASS+exit0+lv) AND claim-type chain -> SIGNED_OFF
    #   5. else -> BLOCKED
    signoff_formula = (
        f'=IF({universal_blank}, "NOT_VERIFIED", '
        f'IF({vs}{R}="NOT_VERIFIED", "NOT_VERIFIED", '
        f'IF({vs}{R}="BLOCKED", "BLOCKED", '
        f'IF({universal_pass}, '
        f'IF({claim_type_chain}, "SIGNED_OFF", "BLOCKED"), '
        f'"BLOCKED"))))'
    )

    # computed_acceptance_status (col 65) — mirrors signoff but maps
    # SIGNED_OFF -> ACCEPTED, BLOCKED -> BLOCKED, NOT_VERIFIED -> PENDING.
    BN = _col(headers, "computed_signoff_status")
    acceptance_formula = (
        f'=IF({BN}{R}="SIGNED_OFF", "ACCEPTED", '
        f'IF({BN}{R}="BLOCKED", "BLOCKED", "PENDING"))'
    )

    # computed_blocking_gap (col 67) — names the first failing required
    # gate, OR the verifier_report_artifact when verifier_status<>PASS.
    BGAP = (
        f'=IF({BN}{R}="SIGNED_OFF", "", '
        f'IF({vs}{R}="BLOCKED", "verifier_status=BLOCKED; see " & {vra}{R}, '
        f'IF({vs}{R}="", "no verifier evidence yet", '
        f'IF({lv}{R}="", "missing last_verified_at_utc", '
        f'"required gate(s) not TRUE for claim_type=" & {G}{R} & '
        f'"; see " & {vra}{R}))))'
    )

    # manual_override_detected (col 68) — TRUE when legacy signoff_status
    # disagrees with computed_signoff_status. Diagnostic only.
    OVERRIDE = (
        f'=IF({AG}{R}="", FALSE, '
        f'IF({AG}{R}<>{BN}{R}, TRUE, FALSE))'
    )

    return {
        "computed_acceptance_status": acceptance_formula,
        "computed_signoff_status":     signoff_formula,
        "computed_blocking_gap":       BGAP,
        "manual_override_detected":    OVERRIDE,
    }


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"FATAL: XLSX not found at {XLSX_PATH}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=False)
    ws = wb["Requirements_Full_Overwrite"]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # Sanity-check schema
    required_cols = (
        "claim_type", "signoff_status",
        "verifier_status", "verifier_exit_code", "last_verified_at_utc",
        "ci_gate_verified", "layer_boundary_verified",
        "required_artifacts_verified", "artifact_payload_hash_verified",
        "runtime_evidence_verified", "evidence_manifest_hash_verified",
        "otel_trace_verified", "source_root_binding_verified",
        "no_bypass_verified", "positive_evidence_verified",
        "replay_receipt_verified", "certifier_signature_verified",
        "expected_fail_reason_verified", "verifier_report_artifact",
        "computed_acceptance_status", "computed_signoff_status",
        "computed_blocking_gap", "manual_override_detected",
    )
    missing = [c for c in required_cols if c not in col_idx]
    if missing:
        print(f"FATAL: schema drift — XLSX missing cols: {missing}", file=sys.stderr)
        return 2

    # Apply formulas to each data row
    n_data_rows = ws.max_row - 1
    for row in range(2, ws.max_row + 1):
        formulas = _build_formulas(headers, row)
        for col_name, formula in formulas.items():
            c = col_idx[col_name]
            ws.cell(row=row, column=c).value = formula

    # Atomic save
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="xlsx_install_", suffix=".xlsx", dir=str(XLSX_PATH.parent))
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        wb.close()
        shutil.move(tmp_path, XLSX_PATH)
    except Exception:
        try: os.unlink(tmp_path)
        except OSError: pass
        raise

    print(f"[install_xlsx_formulas] installed formulas in cols 65-68 across {n_data_rows} rows")
    print(f"  XLSX: {XLSX_PATH}")
    print(f"  Formulas now own: computed_acceptance_status, computed_signoff_status, "
          f"computed_blocking_gap, manual_override_detected")
    print(f"  Formulas reference: claim_type (col {col_idx['claim_type']}) + "
          f"evidence inputs (cols 37-62) + signoff_status (col {col_idx['signoff_status']})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
