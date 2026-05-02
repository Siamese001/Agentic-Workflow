"""Sync the operator CSV → XLSX_FULL_OVERWRITE workbook.

Per operator directive 2026-05-01 15:42 UTC-04:00 — keep the XLSX
workbook in sync with the canonical CSV (operator working surface).

Per operator directive 2026-05-01 15:50 UTC-04:00 (formula-driven
sign-off — plan `runtime-cert-formula-driven-signoff-a8f5c2.md`):
**this sync NEVER writes the formula-owned columns**:
    computed_acceptance_status, computed_signoff_status,
    computed_blocking_gap, manual_override_detected (cols 65-68).
Those cells are owned by the formulas installed via
`install_xlsx_formulas.py` and recompute on Excel open.

What this writes to the XLSX:

  Direct mirror (CSV → XLSX, same column names):
    - signoff_status                (col 33)  — legacy, formula-mirrored
    - signoff_evidence_artifact     (col 34)
    - signoff_evidence_summary      (col 35)
    - signoff_checked_at_utc        (col 36)

  Documentation:
    - authoritative_signoff_source  (col 69)
    - review_notes                  (col 70)

  Dashboard rollup:
    - SIGNED_OFF / BLOCKED / NOT_VERIFIED counts + percent
    - Total rows + drift report
    - Last sync UTC

NEVER touched by this sync:
  - cols 37-62 (evidence inputs — owned by `update_evidence_inputs.py`)
  - cols 65-68 (formula-owned — owned by `install_xlsx_formulas.py`)

Usage:
    python tools/cert/sync_csv_to_xlsx.py
"""

from __future__ import annotations

import csv
import datetime
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

CSV_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened.csv"
)
XLSX_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened_FULL_OVERWRITE.xlsx"
)
RECEIPT_DIR = Path(__file__).resolve().parents[2] / "artifacts" / "certification" / "csv_signoff_updates"


def _utc_now() -> str:
    return datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")


def _extract_blocking_gap(summary: str) -> str:
    """Pull the 'Unblock:' phrase or first sentence from the evidence summary."""
    if not summary:
        return ""
    for marker in ("Unblock:", "Unblocks:", "Unblock at"):
        if marker in summary:
            return summary.split(marker, 1)[1].strip().rstrip(".")[:300]
    # First sentence
    s = summary.split(". ", 1)[0]
    return s[:300]


def _wave_label_for_summary(summary: str) -> str:
    """Extract 'Wave X' prefix from the summary if present."""
    if not summary:
        return ""
    if summary.startswith("Wave"):
        # First "Wave X — title" up to first colon or period
        head = summary.split(":", 1)[0].split(".", 1)[0]
        return head[:80]
    return ""


def main() -> int:
    if not CSV_PATH.exists():
        print(f"FATAL: CSV not found at {CSV_PATH}", file=sys.stderr)
        return 2
    if not XLSX_PATH.exists():
        print(f"FATAL: XLSX not found at {XLSX_PATH}", file=sys.stderr)
        return 2

    csv.field_size_limit(min(sys.maxsize, 2**31 - 1))
    started_at = _utc_now()

    # 1. Read CSV
    with open(CSV_PATH, encoding="utf-8", newline="") as f:
        rdr = csv.DictReader(f)
        csv_rows = list(rdr)
    csv_by_req = {r["req_id"]: r for r in csv_rows}

    # 2. Open XLSX
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=False)
    ws = wb["Requirements_Full_Overwrite"]

    # Build header → 1-based column index
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # Required columns (fail-closed if XLSX schema drifts).
    # NOTE: cols 65-68 (computed_*, manual_override_detected) and col 62
    # (last_verified_at_utc) are deliberately NOT in this list because
    # this sync no longer writes them — they are formula-owned (65-68)
    # or evidence-input owned (62, written only by update_evidence_inputs.py).
    required = (
        "req_id",
        "signoff_status", "signoff_evidence_artifact",
        "signoff_evidence_summary", "signoff_checked_at_utc",
        "authoritative_signoff_source", "review_notes",
    )
    missing = [c for c in required if c not in col_idx]
    if missing:
        print(f"FATAL: XLSX missing required columns: {missing}", file=sys.stderr)
        return 2

    # 3. Index XLSX by req_id (skip header row)
    xlsx_req_to_row = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=col_idx["req_id"]).value
        if rid:
            xlsx_req_to_row[str(rid).strip()] = r

    # 4. Drift report
    csv_only = sorted(set(csv_by_req) - set(xlsx_req_to_row))
    xlsx_only = sorted(set(xlsx_req_to_row) - set(csv_by_req))
    matched = sorted(set(csv_by_req) & set(xlsx_req_to_row))

    # 5. Apply updates
    updated_count = 0
    rollup = {"SIGNED_OFF": 0, "BLOCKED": 0, "NOT_VERIFIED": 0}

    AUTH_SOURCE = (
        "tools/cert/update_csv_signoff.py + per-wave verifiers "
        "(verify_rtc_req_csv_gate, verify_rtc_req_integrated_runtime, "
        "verify_rtc_req_otel_replay, semantic_cache_subclaims composer)"
    )

    for rid in matched:
        csv_row = csv_by_req[rid]
        xlsx_r = xlsx_req_to_row[rid]
        st = csv_row.get("signoff_status", "") or ""
        artifact = csv_row.get("signoff_evidence_artifact", "") or ""
        summary = csv_row.get("signoff_evidence_summary", "") or ""
        ts = csv_row.get("signoff_checked_at_utc", "") or ""

        # Direct mirror cols 33-36 (legacy signoff_* fields)
        ws.cell(row=xlsx_r, column=col_idx["signoff_status"]).value = st
        ws.cell(row=xlsx_r, column=col_idx["signoff_evidence_artifact"]).value = artifact
        ws.cell(row=xlsx_r, column=col_idx["signoff_evidence_summary"]).value = summary
        ws.cell(row=xlsx_r, column=col_idx["signoff_checked_at_utc"]).value = ts

        # NEVER write to:
        #   - col 62 last_verified_at_utc (evidence-input — owned by update_evidence_inputs.py)
        #   - col 65 computed_acceptance_status (formula-owned)
        #   - col 66 computed_signoff_status (formula-owned)
        #   - col 67 computed_blocking_gap (formula-owned)
        #   - col 68 manual_override_detected (formula-owned)

        # Authoritative source (col 69 — documentation)
        ws.cell(row=xlsx_r, column=col_idx["authoritative_signoff_source"]).value = AUTH_SOURCE

        # Review notes — include wave label
        wave = _wave_label_for_summary(summary)
        notes = (
            f"Synced from CSV at {started_at}. "
            f"{('Touched in: ' + wave + '. ') if wave else ''}"
            f"Status: {st}."
        )
        ws.cell(row=xlsx_r, column=col_idx["review_notes"]).value = notes

        rollup[st] = rollup.get(st, 0) + 1
        updated_count += 1

    # 6. Update Dashboard rollup
    if "Dashboard" in wb.sheetnames:
        dash = wb["Dashboard"]
        # Append a fresh rollup block at the bottom of column A/B/C
        next_row = dash.max_row + 2
        dash.cell(row=next_row, column=1).value = f"Sync at {started_at}"
        dash.cell(row=next_row + 1, column=1).value = "SIGNED_OFF"
        dash.cell(row=next_row + 1, column=2).value = rollup["SIGNED_OFF"]
        dash.cell(row=next_row + 1, column=3).value = (
            f"{100 * rollup['SIGNED_OFF'] / max(1, sum(rollup.values())):.1f}%"
        )
        dash.cell(row=next_row + 2, column=1).value = "BLOCKED"
        dash.cell(row=next_row + 2, column=2).value = rollup["BLOCKED"]
        dash.cell(row=next_row + 2, column=3).value = (
            f"{100 * rollup['BLOCKED'] / max(1, sum(rollup.values())):.1f}%"
        )
        dash.cell(row=next_row + 3, column=1).value = "NOT_VERIFIED"
        dash.cell(row=next_row + 3, column=2).value = rollup["NOT_VERIFIED"]
        dash.cell(row=next_row + 3, column=3).value = (
            f"{100 * rollup['NOT_VERIFIED'] / max(1, sum(rollup.values())):.1f}%"
        )
        dash.cell(row=next_row + 4, column=1).value = "TOTAL"
        dash.cell(row=next_row + 4, column=2).value = sum(rollup.values())
        dash.cell(row=next_row + 5, column=1).value = "Sync source"
        dash.cell(row=next_row + 5, column=2).value = str(CSV_PATH)
        dash.cell(row=next_row + 6, column=1).value = "Drift (CSV-only)"
        dash.cell(row=next_row + 6, column=2).value = ", ".join(csv_only) or "(none)"
        dash.cell(row=next_row + 7, column=1).value = "Drift (XLSX-only)"
        dash.cell(row=next_row + 7, column=2).value = ", ".join(xlsx_only) or "(none)"

    # 7. Atomic save
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="xlsx_sync_", suffix=".xlsx", dir=str(XLSX_PATH.parent))
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        wb.close()
        shutil.move(tmp_path, XLSX_PATH)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise

    # 8. Receipt
    RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
    receipt = {
        "tool": "tools/cert/sync_csv_to_xlsx.py",
        "executed_at_utc": started_at,
        "csv_path": str(CSV_PATH),
        "xlsx_path": str(XLSX_PATH),
        "matched_count": len(matched),
        "csv_only": csv_only,
        "xlsx_only": xlsx_only,
        "updated_rows": updated_count,
        "post_sync_rollup": rollup,
        "post_sync_total": sum(rollup.values()),
    }
    receipt_path = RECEIPT_DIR / f"{started_at.replace(':', '-')}_xlsx_sync.json"
    receipt_path.write_text(
        json.dumps(receipt, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    print(f"[sync_csv_to_xlsx] CSV → XLSX sync complete")
    print(f"  CSV : {CSV_PATH}")
    print(f"  XLSX: {XLSX_PATH}")
    print(f"  matched: {len(matched)} req_ids")
    if csv_only:
        print(f"  CSV-only (missing from XLSX): {csv_only}")
    if xlsx_only:
        print(f"  XLSX-only (missing from CSV): {xlsx_only}")
    print(f"  rollup: {rollup}  total={sum(rollup.values())}")
    print(f"  receipt: {receipt_path.relative_to(receipt_path.parents[3])}")
    return 0 if (not csv_only and not xlsx_only) else 1


if __name__ == "__main__":
    sys.exit(main())
