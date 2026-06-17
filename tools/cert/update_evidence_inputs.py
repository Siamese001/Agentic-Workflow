"""F0.4 — Update evidence-input fields (cols 37-62) for one or more reqs.

This is the **only** path Codex or operators use to populate the
26 evidence-input fields per row going forward. Replaces direct edits
to `signoff_status` (which is now a legacy mirror of the formula-owned
`computed_signoff_status`).

Per plan `runtime-cert-formula-driven-signoff-a8f5c2.md` §F0.4.

Hard rules enforced:
  - Refuses to write any of the 4 formula-owned columns (cols 65-68).
  - Refuses to write outside the declared 26 evidence-input columns.
  - Atomic save (temp file + rename).
  - Receipt JSON written to artifacts/certification/csv_signoff_updates/.

Usage:

  # Apply a single set of evidence inputs to a list of req_ids
  python tools/cert/update_evidence_inputs.py \
      --req-ids "RTC-REQ-001,RTC-REQ-002" \
      --evidence-json '{"verifier_status":"PASS","verifier_exit_code":0,"ci_gate_verified":true,"last_verified_at_utc":"2026-05-01T20:00:00+00:00"}' \
      --wave-label "F1 — matrix governance backfill"

  # Or load evidence from a JSON file mapping req_id -> {field: value, ...}
  python tools/cert/update_evidence_inputs.py \
      --batch-json artifacts/certification/F1_evidence_inputs.json \
      --wave-label "F1 — matrix governance backfill"
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

import openpyxl

XLSX_PATH = Path(
    r"C:\Users\amita\Downloads\runtime_certification_requirements_100_percent_hardened_FULL_OVERWRITE.xlsx"
)
RECEIPT_DIR = Path(__file__).resolve().parents[2] / "artifacts" / "certification" / "csv_signoff_updates"

EVIDENCE_INPUT_COLUMNS = (
    "verifier_status",
    "verifier_exit_code",
    "verifier_report_artifact",
    "verifier_report_sha256",
    "evidence_manifest_artifact",
    "evidence_manifest_sha256",
    "evidence_manifest_hash_verified",
    "required_artifacts_verified",
    "positive_evidence_verified",
    "negative_controls_verified",
    "expected_fail_reason_verified",
    "ci_gate_verified",
    "runtime_evidence_verified",
    "otel_trace_verified",
    "replay_receipt_verified",
    "no_bypass_verified",
    "uwg_write_path_verified",
    "layer_boundary_verified",
    "source_root_binding_verified",
    "artifact_payload_hash_verified",
    "merkle_leaf_verified",
    "proof_depth_verified",
    "certifier_identity",
    "certifier_signature_artifact",
    "certifier_signature_verified",
    "last_verified_at_utc",
)
ALLOWED_FIELDS = set(EVIDENCE_INPUT_COLUMNS)

FORMULA_OWNED_COLUMNS = (
    "computed_acceptance_status",
    "computed_signoff_status",
    "computed_blocking_gap",
    "manual_override_detected",
)
FORBIDDEN_FIELDS = set(FORMULA_OWNED_COLUMNS)


def _utc_now() -> str:
    return datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")


def _validate_evidence(evidence: dict[str, Any]) -> tuple[bool, list[str]]:
    """Validate that ``evidence`` only references allowed evidence-input fields."""
    errors: list[str] = []
    for k in evidence:
        if k in FORBIDDEN_FIELDS:
            errors.append(f"FORBIDDEN: {k!r} is formula-owned; remove from evidence input")
        elif k not in ALLOWED_FIELDS:
            errors.append(f"UNKNOWN: {k!r} is not one of the 26 evidence-input fields")
    return (len(errors) == 0, errors)


def _coerce_bool(value: Any) -> Any:
    """openpyxl writes Python bool True/False as Excel TRUE/FALSE.
    Strings 'true'/'false' are also accepted (case-insensitive)."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        s = value.strip().lower()
        if s == "true":
            return True
        if s == "false":
            return False
    return value


def _file_sha256(path: str | Path) -> str:
    """Compute sha256 of a file. Returns '' if file missing."""
    p = Path(path)
    if not p.exists():
        return ""
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _autocompute_hashes(evidence: dict[str, Any]) -> dict[str, Any]:
    """If verifier_report_artifact / evidence_manifest_artifact are
    paths and their *_sha256 partners are absent, auto-fill them from
    the file content. This is convenience for callers."""
    out = dict(evidence)
    pair = (
        ("verifier_report_artifact", "verifier_report_sha256"),
        ("evidence_manifest_artifact", "evidence_manifest_sha256"),
    )
    for art_key, sha_key in pair:
        art = out.get(art_key)
        if art and not out.get(sha_key):
            try:
                # Allow paths relative to repo root
                p = Path(art)
                if not p.is_absolute():
                    repo_root = Path(__file__).resolve().parents[2]
                    p = repo_root / p
                computed = _file_sha256(p)
                if computed:
                    out[sha_key] = computed
            except (OSError, ValueError):
                pass
    return out


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--req-ids", help="Comma-separated req_ids to update")
    p.add_argument("--evidence-json", help="JSON string of {field: value} to apply to all --req-ids")
    p.add_argument("--batch-json",
                   help="Path to JSON file mapping req_id -> {field: value, ...}")
    p.add_argument("--wave-label", required=True,
                   help="Short wave label for the receipt + review_notes")
    p.add_argument("--default-last-verified", default=None,
                   help="If set, fill last_verified_at_utc with this UTC timestamp on every row that omits it")
    return p.parse_args()


def main() -> int:
    args = _parse_args()
    started_at = _utc_now()

    # 1. Build the per-req evidence map
    per_req_evidence: dict[str, dict[str, Any]] = {}

    if args.batch_json:
        bp = Path(args.batch_json)
        if not bp.exists():
            print(f"FATAL: batch JSON not found: {bp}", file=sys.stderr)
            return 2
        loaded = json.loads(bp.read_text(encoding="utf-8"))
        if not isinstance(loaded, dict):
            print("FATAL: batch JSON must be a dict mapping req_id -> evidence dict", file=sys.stderr)
            return 2
        per_req_evidence.update(loaded)

    if args.evidence_json:
        if not args.req_ids:
            print("FATAL: --evidence-json requires --req-ids", file=sys.stderr)
            return 2
        ev = json.loads(args.evidence_json)
        rids = [r.strip() for r in args.req_ids.split(",") if r.strip()]
        for rid in rids:
            base = per_req_evidence.get(rid, {})
            base.update(ev)
            per_req_evidence[rid] = base

    if not per_req_evidence:
        print("FATAL: no req-id/evidence pairs to write", file=sys.stderr)
        return 2

    # Default last_verified_at_utc
    default_lv = args.default_last_verified or started_at
    for rid, ev in per_req_evidence.items():
        if "last_verified_at_utc" not in ev:
            ev["last_verified_at_utc"] = default_lv

    # 2. Validate
    all_errors: list[str] = []
    for rid, ev in per_req_evidence.items():
        ok, errs = _validate_evidence(ev)
        if not ok:
            for e in errs:
                all_errors.append(f"{rid}: {e}")
    if all_errors:
        print("FATAL: evidence-input validation failed:", file=sys.stderr)
        for e in all_errors:
            print(f"  {e}", file=sys.stderr)
        return 2

    # 3. Auto-compute report sha256 where convenient
    for rid, ev in per_req_evidence.items():
        per_req_evidence[rid] = _autocompute_hashes(ev)

    # 4. Open XLSX, locate rows
    if not XLSX_PATH.exists():
        print(f"FATAL: XLSX not found at {XLSX_PATH}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=False)
    ws = wb["Requirements_Full_Overwrite"]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # Schema check
    missing_cols = [c for c in EVIDENCE_INPUT_COLUMNS if c not in col_idx]
    if missing_cols:
        print(f"FATAL: XLSX missing evidence-input columns: {missing_cols}", file=sys.stderr)
        return 2

    # Index rows by req_id
    req_to_row: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=col_idx["req_id"]).value
        if rid:
            req_to_row[str(rid).strip()] = r

    not_found = sorted(set(per_req_evidence) - set(req_to_row))
    if not_found:
        print(f"FATAL: req_ids not found in XLSX: {not_found}", file=sys.stderr)
        return 2

    # 5. Apply evidence-input writes
    n_cells_written = 0
    for rid, ev in per_req_evidence.items():
        r = req_to_row[rid]
        for field, value in ev.items():
            c = col_idx[field]
            coerced = _coerce_bool(value)
            ws.cell(row=r, column=c).value = coerced
            n_cells_written += 1

    # 6. Atomic save
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="xlsx_evidence_", suffix=".xlsx", dir=str(XLSX_PATH.parent))
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        wb.close()
        shutil.move(tmp_path, XLSX_PATH)
    except Exception:
        try: os.unlink(tmp_path)
        except OSError: pass
        raise

    # 7. Receipt
    RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
    receipt = {
        "tool": "tools/cert/update_evidence_inputs.py",
        "wave_label": args.wave_label,
        "executed_at_utc": started_at,
        "xlsx_path": str(XLSX_PATH),
        "req_ids": sorted(per_req_evidence),
        "evidence_per_req": per_req_evidence,
        "n_reqs": len(per_req_evidence),
        "n_cells_written": n_cells_written,
    }
    safe_label = args.wave_label.replace(" ", "_").replace("/", "-").replace(":", "-")
    safe_label = "".join(c for c in safe_label if c.isalnum() or c in "_-.")
    receipt_path = RECEIPT_DIR / f"{started_at.replace(':', '-')}_evidence_{safe_label}.json"
    receipt_path.write_text(
        json.dumps(receipt, indent=2, sort_keys=True, default=str) + "\n",
        encoding="utf-8",
    )

    print(f"[update_evidence_inputs] {args.wave_label}")
    print(f"  XLSX: {XLSX_PATH}")
    print(f"  reqs: {len(per_req_evidence)}  cells: {n_cells_written}")
    print(f"  receipt: {receipt_path.relative_to(receipt_path.parents[3])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
